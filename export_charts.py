#!/usr/bin/env python3
"""
LibreChat 图表数据导出工具

将 AI 生成的图表数据 JSON 转换为 Excel 和 PDF 文件

使用方法：
    python export_charts.py chart_data.json

依赖安装：
    pip install pandas openpyxl matplotlib

作者：LibreChat Contract Analysis Extension
"""

import json
import sys
import os
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages
    plt.rcParams['font.family'] = ['DejaVu Sans', 'Arial Unicode MS', 'sans-serif']
except ImportError as e:
    print(f"❌ 缺少必需的库: {e}")
    print("\n请安装依赖:")
    print("pip install pandas openpyxl matplotlib")
    sys.exit(1)


def load_chart_data(json_file):
    """加载图表数据 JSON 文件"""
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data
    except FileNotFoundError:
        print(f"❌ 文件未找到: {json_file}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"❌ JSON 解析错误: {e}")
        sys.exit(1)


def validate_chart_data(data):
    """验证图表数据结构"""
    if not isinstance(data, dict):
        print("❌ 数据格式错误：应为字典类型")
        return False

    if not data.get('has_data', False):
        print("⚠️  数据中没有图表信息")
        return False

    if 'charts' not in data or not isinstance(data['charts'], list):
        print("❌ 数据格式错误：缺少 charts 数组")
        return False

    # 验证每个图表
    for i, chart in enumerate(data['charts'], 1):
        if not all(k in chart for k in ['chart_title', 'chart_type', 'data']):
            print(f"❌ 图表 {i} 缺少必需字段")
            return False

        if not isinstance(chart['data'], list) or len(chart['data']) == 0:
            print(f"❌ 图表 {i} 数据为空")
            return False

        for item in chart['data']:
            if 'label' not in item or 'value' not in item:
                print(f"❌ 图表 {i} 数据项缺少 label 或 value")
                return False

    return True


def export_to_excel(data, output_file):
    """导出为 Excel 文件"""
    print(f"\n📊 开始生成 Excel 文件...")

    charts = data['charts']

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 1. 创建摘要工作表
        summary_data = {
            '图表编号': [],
            '标题': [],
            '类型': [],
            '数据点数': [],
            '来源页': [],
            '说明': []
        }

        for i, chart in enumerate(charts, 1):
            summary_data['图表编号'].append(i)
            summary_data['标题'].append(chart['chart_title'])
            summary_data['类型'].append(chart['chart_type'])
            summary_data['数据点数'].append(len(chart['data']))

            page_info = chart.get('page_number', 'N/A')
            if page_info and page_info != 'N/A':
                summary_data['来源页'].append(f"第 {page_info} 页")
            else:
                summary_data['来源页'].append(chart.get('category', 'N/A'))

            summary_data['说明'].append(chart.get('explanation', ''))

        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='摘要', index=False)

        # 2. 为每个图表创建工作表
        for i, chart in enumerate(charts, 1):
            # 创建图表数据框
            chart_df = pd.DataFrame(chart['data'])

            # 重命名列为中文
            if 'label' in chart_df.columns and 'value' in chart_df.columns:
                chart_df.rename(columns={'label': '标签', 'value': '数值'}, inplace=True)

            # 工作表名称（限制 31 个字符）
            sheet_name = f"图表{i}"
            title_len = len(chart['chart_title'])
            if title_len <= 20:
                sheet_name = chart['chart_title'][:31]
            elif title_len <= 25:
                sheet_name = f"图{i}-{chart['chart_title'][:25]}"

            # 写入图表数据
            chart_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # 添加元数据
            metadata_df = pd.DataFrame({
                '属性': ['图表标题', '图表类型', '来源页', '说明'],
                '值': [
                    chart['chart_title'],
                    chart['chart_type'],
                    f"第 {chart.get('page_number', 'N/A')} 页" if chart.get('page_number') and chart['page_number'] != 'N/A' else chart.get('category', 'N/A'),
                    chart.get('explanation', '')
                ]
            })

            # 在数据下方写入元数据（留 2 行间隔）
            start_row = len(chart_df) + 3
            metadata_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)

    # 格式化 Excel（可选）
    try:
        wb = load_workbook(output_file)

        # 格式化摘要工作表
        if '摘要' in wb.sheetnames:
            ws = wb['摘要']
            # 标题行加粗
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_file)
    except Exception as e:
        print(f"⚠️  Excel 格式化时出现警告: {e}")

    print(f"✅ Excel 文件已生成: {output_file}")


def export_to_pdf(data, output_file):
    """导出为 PDF 文件"""
    print(f"\n📄 开始生成 PDF 文件...")

    charts = data['charts']

    with PdfPages(output_file) as pdf:
        # 第 1 页：摘要页
        fig = plt.figure(figsize=(8.5, 11))
        ax = fig.add_subplot(111)
        ax.axis('off')

        title_text = "合同数据分析报告"
        ax.text(0.5, 0.95, title_text, ha='center', va='top',
                fontsize=18, fontweight='bold', transform=ax.transAxes)

        summary_text = f"共生成图表: {len(charts)}\n\n"
        summary_text += "图表列表:\n"
        for i, chart in enumerate(charts, 1):
            if chart.get('page_number') and chart['page_number'] != 'N/A':
                summary_text += f"{i}. 第 {chart['page_number']} 页: {chart['chart_title']}\n"
            else:
                summary_text += f"{i}. {chart['chart_title']} ({chart.get('category', '通用')})\n"

        ax.text(0.1, 0.85, summary_text, ha='left', va='top',
                fontsize=11, transform=ax.transAxes, family='sans-serif')

        pdf.savefig(fig, bbox_inches='tight')
        plt.close(fig)

        # 第 2-N 页：每个图表一页
        for i, chart in enumerate(charts, 1):
            fig, ax = plt.subplots(figsize=(10, 7))

            labels = [item['label'] for item in chart['data']]
            values = [item['value'] for item in chart['data']]
            chart_type = chart['chart_type']

            # 绘制图表
            if chart_type == 'bar':
                bars = ax.bar(labels, values, color='steelblue', alpha=0.8, edgecolor='navy')
                # 在柱子上添加数值标签
                for bar in bars:
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width()/2., height,
                            f'{height:.0f}',
                            ha='center', va='bottom', fontsize=9)
            else:  # line
                ax.plot(labels, values, marker='o', linewidth=2.5,
                        markersize=8, color='steelblue')
                # 在点上添加数值标签
                for j, (label, value) in enumerate(zip(labels, values)):
                    ax.text(j, value, f'{value:.0f}',
                            ha='center', va='bottom', fontsize=9)

            # 标题（含来源页码）
            if chart.get('page_number') and chart['page_number'] != 'N/A':
                title = f"{chart['chart_title']}\n(来源: 第 {chart['page_number']} 页)"
            else:
                title = chart['chart_title']

            ax.set_title(title, fontsize=14, fontweight='bold', pad=20)

            ax.set_xlabel('类别', fontsize=11, fontweight='bold')
            ax.set_ylabel('数值', fontsize=11, fontweight='bold')
            ax.grid(True, alpha=0.3, linestyle='--')

            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()

            # 添加解释文本（底部）
            if chart.get('explanation'):
                fig.text(0.5, 0.02, chart['explanation'],
                         ha='center', fontsize=9, style='italic',
                         wrap=True, color='gray')

            pdf.savefig(fig, bbox_inches='tight')
            plt.close(fig)

    print(f"✅ PDF 文件已生成: {output_file}")


def main():
    """主函数"""
    print("=" * 60)
    print("LibreChat 图表数据导出工具")
    print("=" * 60)

    # 检查命令行参数
    if len(sys.argv) < 2:
        print("\n使用方法:")
        print(f"  python {sys.argv[0]} <chart_data.json>")
        print("\n示例:")
        print(f"  python {sys.argv[0]} my_charts.json")
        sys.exit(1)

    json_file = sys.argv[1]

    # 加载数据
    print(f"\n📂 加载数据文件: {json_file}")
    data = load_chart_data(json_file)

    # 验证数据
    print("🔍 验证数据结构...")
    if not validate_chart_data(data):
        sys.exit(1)

    print(f"✅ 发现 {len(data['charts'])} 个图表")

    # 生成输出文件名
    base_name = Path(json_file).stem
    excel_file = f"{base_name}.xlsx"
    pdf_file = f"{base_name}.pdf"

    # 导出 Excel
    try:
        export_to_excel(data, excel_file)
    except Exception as e:
        print(f"❌ Excel 导出失败: {e}")

    # 导出 PDF
    try:
        export_to_pdf(data, pdf_file)
    except Exception as e:
        print(f"❌ PDF 导出失败: {e}")

    print("\n" + "=" * 60)
    print("✨ 导出完成！")
    print("=" * 60)
    print(f"\n生成的文件:")
    if os.path.exists(excel_file):
        print(f"  📊 Excel: {excel_file}")
    if os.path.exists(pdf_file):
        print(f"  📄 PDF: {pdf_file}")
    print()


if __name__ == "__main__":
    main()
